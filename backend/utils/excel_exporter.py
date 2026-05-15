import os
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font
from openpyxl.utils.dataframe import dataframe_to_rows
from typing import Dict, Any, List

def _auto_adjust_columns(worksheet):
    """Utility to auto-fit column widths."""
    for col in worksheet.columns:
        max_length = 0
        column = col[0].column_letter
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        worksheet.column_dimensions[column].width = min(max_length + 2, 80)

def export_to_excel(extracted_kpis: Dict[str, Any], 
                    audit_trail: List[Dict[str, Any]], 
                    peer_multiples_df: pd.DataFrame, 
                    valuation_ranges: Dict[str, tuple], 
                    output_path: str,
                    book_value: float = None) -> str:
    """
    Generates a 4-sheet formatted Excel deliverable with KPI Confidence, Audit Trail, 
    Peer Comps, and Valuation Ranges.
    """
    wb = Workbook()
    
    # ------------------ STYLES ------------------
    bold_font = Font(bold=True)
    # Highlight colors
    high_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid") # Pastel Green
    med_fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")  # Pastel Yellow
    low_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")  # Pastel Red
    blue_fill = PatternFill(start_color="DCE6F1", end_color="DCE6F1", fill_type="solid") # Pastel Blue
    
    # ==========================================
    # Sheet 1: KPIs
    # ==========================================
    ws1 = wb.active
    ws1.title = "KPIs"
    
    ws1.append(["KPI Name", "Value"])
    ws1.cell(row=1, column=1).font = bold_font
    ws1.cell(row=1, column=2).font = bold_font
    
    row_idx = 2
    for kpi, data in extracted_kpis.items():
        val = data.get("value")
        conf = str(data.get("confidence", "")).upper()
        
        # Capture book value if it exists and wasn't explicitly passed
        if book_value is None and "book" in kpi.lower():
            book_value = val
            
        ws1.append([kpi, val])
        
        # Apply color coding
        fill = None
        if conf == "HIGH":
            fill = high_fill
        elif conf == "MEDIUM":
            fill = med_fill
        elif conf == "LOW":
            fill = low_fill
            
        if fill:
            ws1.cell(row=row_idx, column=1).fill = fill
            ws1.cell(row=row_idx, column=2).fill = fill
            
        row_idx += 1
        
    _auto_adjust_columns(ws1)

    # ==========================================
    # Sheet 2: Audit Trail
    # ==========================================
    ws2 = wb.create_sheet(title="Audit Trail")
    headers = ["KPI", "Value", "Confidence", "Page", "Line Text", "Source"]
    ws2.append(headers)
    for i in range(1, len(headers) + 1):
        ws2.cell(row=1, column=i).font = bold_font
        
    for item in audit_trail:
        ws2.append([
            item.get("kpi"),
            item.get("value"),
            item.get("confidence"),
            item.get("page_number"),
            item.get("line_text"),
            item.get("source")
        ])
        
    _auto_adjust_columns(ws2)

    # ==========================================
    # Sheet 3: Comps
    # ==========================================
    ws3 = wb.create_sheet(title="Comps")
    
    # Write DataFrame to worksheet
    for r in dataframe_to_rows(peer_multiples_df, index=False, header=True):
        ws3.append(r)
        
    # Bold the header row
    for cell in ws3[1]:
        cell.font = bold_font
        
    # Calculate medians for the multiple columns
    target_cols = ['ev_ebitda', 'ev_revenue', 'pe_ratio', 'p_book']
    median_row = ["Median"] + [""] * (len(peer_multiples_df.columns) - 1)
    
    for idx, col in enumerate(peer_multiples_df.columns):
        if col in target_cols:
            valid_data = peer_multiples_df[col].dropna()
            valid_data = valid_data[valid_data > 0]
            if not valid_data.empty:
                median_row[idx] = valid_data.median()
                
    ws3.append(median_row)
    
    # Highlight final median row in blue and bold it
    max_row = ws3.max_row
    for cell in ws3[max_row]:
        cell.fill = blue_fill
        cell.font = bold_font
        
    _auto_adjust_columns(ws3)

    # ==========================================
    # Sheet 4: Valuation
    # ==========================================
    ws4 = wb.create_sheet(title="Valuation")
    headers4 = ["Methodology", "Low", "High", "Midpoint"]
    ws4.append(headers4)
    for i in range(1, 5):
        ws4.cell(row=1, column=i).font = bold_font
        
    for meth, (low, high) in valuation_ranges.items():
        mid = (low + high) / 2.0
        ws4.append([meth, low, high, mid])
        
    # Add the final Book Value row
    final_bv = book_value if book_value is not None else "N/A"
    ws4.append(["Book Value", final_bv, final_bv, final_bv])
    
    _auto_adjust_columns(ws4)

    # ==========================================
    # Save and Return
    # ==========================================
    os.makedirs(os.path.dirname(os.path.abspath(output_path)), exist_ok=True)
    wb.save(output_path)
    return output_path
